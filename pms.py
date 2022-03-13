from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
import time
import openpyxl
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

path = " " #Excel file path

wb = openpyxl.load_workbook(path)

sheet = wb.active

browser = webdriver.Chrome(executable_path="C:/Users/Bharath/Desktop/chromedriver.exe")

browser.get("http://samagrashiksha.in/")
time.sleep(2)
browser.find_element(By.ID, "txtusername").send_keys("")
time.sleep(2)
browser.find_element(By.ID, "txtpwd").send_keys("")
time.sleep(5)
browser.find_element(By.ID, "txtTuring").send_keys(input(int()))
time.sleep(10)
browser.find_element(By.ID, "Button1").click()
time.sleep(1)

browser.get("http://samagrashiksha.in/Block/cwsn_boys_dist.aspx")
time.sleep(5)
select = Select(browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_ddlFinancialYear"))
select.select_by_visible_text("2021-2022")
time.sleep(5)

for i in range(2, sheet.max_row+1):
    cel_dise = sheet.cell(row=i,column=1)
    browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_txtudise").send_keys(cel_dise.value)
    time.sleep(1)
    browser.find_element(By.ID, "Dashboard_main").click()
    time.sleep(5)
    st_n = sheet.cell(row=i,column=2)
    st_f = sheet.cell(row=i,column=3)
    st_m = sheet.cell(row=i,column=4)
    st_a = sheet.cell(row=i,column=5)
    st_p = sheet.cell(row=i,column=6)
    st_ad = sheet.cell(row=i, column=7)
    st_dob = sheet.cell(row=i, column=8)
    st_cls = sheet.cell(row=i, column=9)
    st_db = sheet.cell(row=i, column=10)
    st_ac = sheet.cell(row=i, column=11)
    st_ifsc = sheet.cell(row=i, column=12)
    st_bank = sheet.cell(row=i, column=13)


    browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_Gridview1_ctl02_txtstuname").send_keys(st_n.value)
    time.sleep(1)
    browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_Gridview1_ctl02_txtfather").send_keys(st_f.value)
    time.sleep(1)
    browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_Gridview1_ctl02_txtmother").send_keys(st_m.value)
    time.sleep(1)
    browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_Gridview1_ctl02_txtaddress").send_keys(st_a.value)
    time.sleep(1)
    browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_Gridview1_ctl02_txtpin").send_keys(st_p.value)
    time.sleep(1)
    if st_ad.value == None:
        browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_Gridview1_ctl02_txtadhar").send_keys("")
    else:
        browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_Gridview1_ctl02_txtadhar").send_keys(st_ad.value)
    date = (str(st_dob.value))
    ac = date[0:10]
    red = ac.split("-")
    # red = red.reverse()
    f_date = (str(red[2] + "/" + red[1] + "/" + red[0]))
    browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_Gridview1_ctl02_txtdob").send_keys(f_date)
    time.sleep(1)

    cla_sel = Select(browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_Gridview1_ctl02_ddlcls"))
    if st_cls.value == "1":
        cla_sel.select_by_visible_text("1th")
    elif st_cls.value == "2":
        cla_sel.select_by_visible_text("2th")
    elif st_cls.value == "3":
        cla_sel.select_by_visible_text("3th")
    elif st_cls.value == "4":
        cla_sel.select_by_visible_text("4th")
    elif st_cls.value == "5":
        cla_sel.select_by_visible_text("5th")
    elif st_cls.value == "6":
        cla_sel.select_by_visible_text("6th")
    elif st_cls.value == "7":
        cla_sel.select_by_visible_text("7th")
    elif st_cls.value == "8":
        cla_sel.select_by_visible_text("8th")
    elif st_cls.value == "9":
        cla_sel.select_by_visible_text("9th")
    elif st_cls.value == "10":
        cla_sel.select_by_visible_text("10th")
    else:
        cla_sel.select_by_visible_text("NA")

    st_typ = Select(browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_Gridview1_ctl02_ddltype"))
    st_typ.select_by_visible_text("Day Scholar")
    time.sleep(1)
    st_dis = Select(browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_Gridview1_ctl02_ddldisable"))
    if st_db.value == "Visual Impairment (Low-vision)":
        st_dis.select_by_visible_text("Low Vision")
    elif st_db.value == "Speech and Language Disability":
        st_dis.select_by_visible_text("Speech and Language")
    elif st_db.value == "Intellectual Disability":
        st_dis.select_by_visible_text("Intellectual Disability")
    elif st_db.value == "Loco motor impairment":
        st_dis.select_by_visible_text("Locomotor Disability")
    elif st_db.value == "Multiple Disabilities including Deaf Blindness":
        st_dis.select_by_visible_text("Multiple Disability incl. DB")
    elif st_db.value == "Multiple Sclerosis":
        st_dis.select_by_visible_text("Multiple Sclerosis")
    elif st_db.value == "Muscular Dystrophy":
        st_dis.select_by_visible_text("Muscular Dystrophy")
    elif st_db.value == "Hearing Imparement (Deaf and Hard of Hearing":
        st_dis.select_by_visible_text("Hearing Impaired (deaf & HOH)")
    elif st_db.value == "Cerebral Palsy":
        st_dis.select_by_visible_text("Cerebal Palsy")
    elif st_db.value == "Visual Impairment (Blindness)":
        st_dis.select_by_visible_text("Blindness")
    elif st_db.value == "Specific Learning Disability":
        st_dis.select_by_visible_text("Specific Learning Disability")
    elif st_db.value == "HEARING IMPAREMENT":
        st_dis.select_by_visible_text("Hearing Impaired (deaf & HOH)")
    elif st_db.value == "Thalassemia":
        st_dis.select_by_visible_text("Thalassemia")
    elif st_db.value == "Chronic Neurological Conditios":
        st_dis.select_by_visible_text("Thalassemia")
    elif st_db.value == "Chronic Neurological Conditios":
        st_dis.select_by_visible_text("Thalassemia")
    elif st_db.value == "Parkinson's Disease":
        st_dis.select_by_visible_text("Parkinson's disease")
    else:
        st_dis.select_by_visible_text("Select")

    time.sleep(1)
    browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_Gridview1_ctl02_txtaccount").send_keys(st_ac.value)
    time.sleep(1)
    browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_Gridview1_ctl02_txtifsc").send_keys(st_ifsc.value)
    time.sleep(1)
    ad_seed = Select(browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_Gridview1_ctl02_ddlseed"))
    ad_seed.select_by_visible_text("Y")
    time.sleep(1)
    amt_trn = Select(browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_Gridview1_ctl02_ddlmtransfer"))
    amt_trn.select_by_visible_text("Electronic Fund transfer, directly through Bank")
    time.sleep(1)
    browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_Gridview1_ctl02_txtbanknm").send_keys(st_bank.value)
    time.sleep(2)
    browser.find_element(By.ID, "ctl00_ContentPlaceHolder1_Submit").click()
    time.sleep(1)

    WebDriverWait(browser, 5).until(EC.alert_is_present())
    browser.switch_to.alert.accept()
    sheet.cell(row=i, column=14).value = "Entry Done"
    wb.save("C:/Users/Bharath/Desktop/pfms entry.xlsx")
